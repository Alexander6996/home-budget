from datetime import datetime
from io import BytesIO

from flask import jsonify, redirect, render_template, request, send_file, session, url_for, flash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from auth_utils import login_required
from database import get_db_connection


def register_transaction_routes(app):
    @app.route('/add', methods=['GET', 'POST'])
    @login_required
    def add_transaction():
        """Добавление новой транзакции"""
        with get_db_connection() as conn:
            categories = conn.execute('SELECT * FROM categories').fetchall()

            if request.method == 'POST':
                type_ = request.form['type']
                amount = float(request.form['amount'])
                category_id = request.form.get('category_id')
                description = request.form.get('description', '')
                date = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))

                if not category_id:
                    flash('Выберите категорию!', 'error')
                    return redirect(url_for('add_transaction'))

                conn.execute('''
                    INSERT INTO transactions
                    (type, amount, category_id, description, date, user_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (type_, amount, category_id, description, date, session['user_id']))
                conn.commit()

                flash('Транзакция успешно добавлена!', 'success')
                return redirect(url_for('index'))

        return render_template('add_transaction.html', categories=categories)

    @app.route('/transactions')
    @login_required
    def view_transactions():
        """Просмотр всех транзакций"""
        with get_db_connection() as conn:
            filter_type = request.args.get('type', 'all')
            filter_category = request.args.get('category', 'all')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            search = request.args.get('search', '').strip()

            query = '''
                SELECT t.*, c.name as category_name, c.icon as category_icon
                FROM transactions t
                LEFT JOIN categories c ON t.category_id = c.id
                WHERE t.user_id = ?
            '''
            params = [session['user_id']]

            if filter_type != 'all':
                query += ' AND t.type = ?'
                params.append(filter_type)

            if filter_category != 'all':
                query += ' AND t.category_id = ?'
                params.append(filter_category)

            if start_date:
                query += ' AND t.date >= ?'
                params.append(start_date)

            if end_date:
                query += ' AND t.date <= ?'
                params.append(end_date)

            query += ' ORDER BY t.date DESC, t.created_at DESC'

            transactions = conn.execute(query, params).fetchall()
            categories = conn.execute('SELECT * FROM categories').fetchall()

        if search:
            search_value = search.casefold()
            filtered_transactions = []

            for trans in transactions:
                description = (trans['description'] or '').casefold()
                category_name = (trans['category_name'] or '').casefold()

                if search_value in description or search_value in category_name:
                    filtered_transactions.append(trans)

            transactions = filtered_transactions

        return render_template(
            'transactions.html',
            transactions=transactions,
            categories=categories,
            filter_type=filter_type,
            filter_category=filter_category,
            start_date=start_date,
            end_date=end_date,
            search=search
        )

    @app.route('/edit/<int:transaction_id>', methods=['GET', 'POST'])
    @login_required
    def edit_transaction(transaction_id):
        """Редактирование транзакции"""
        with get_db_connection() as conn:
            transaction = conn.execute(
                'SELECT * FROM transactions WHERE id = ? AND user_id = ?',
                (transaction_id, session['user_id'])
            ).fetchone()

            if not transaction:
                flash('Транзакция не найдена', 'error')
                return redirect(url_for('view_transactions'))

            categories = conn.execute('SELECT * FROM categories').fetchall()

            if request.method == 'POST':
                type_ = request.form['type']
                amount = float(request.form['amount'])
                category_id = request.form.get('category_id')
                description = request.form.get('description', '')
                date = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))

                conn.execute('''
                    UPDATE transactions
                    SET type = ?, amount = ?, category_id = ?, description = ?, date = ?
                    WHERE id = ? AND user_id = ?
                ''', (type_, amount, category_id, description, date, transaction_id, session['user_id']))
                conn.commit()

                flash('Транзакция успешно обновлена!', 'success')
                return redirect(url_for('view_transactions'))

        return render_template(
            'edit_transaction.html',
            transaction=transaction,
            categories=categories
        )

    @app.route('/delete/<int:transaction_id>')
    @login_required
    def delete_transaction(transaction_id):
        """Удаление транзакции"""
        with get_db_connection() as conn:
            transaction = conn.execute(
                'SELECT * FROM transactions WHERE id = ? AND user_id = ?',
                (transaction_id, session['user_id'])
            ).fetchone()

            if transaction:
                conn.execute(
                    'DELETE FROM transactions WHERE id = ? AND user_id = ?',
                    (transaction_id, session['user_id'])
                )
                conn.commit()
                flash('Транзакция успешно удалена!', 'success')
            else:
                flash('Транзакция не найдена', 'error')

        return redirect(url_for('view_transactions'))

    @app.route('/export/excel')
    @login_required
    def export_excel():
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        filter_type = request.args.get('type', 'all')

        query = '''
            SELECT t.date, t.type, t.amount, c.name as category, t.description
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.id
            WHERE t.user_id = ?
        '''
        params = [session['user_id']]

        if start_date:
            query += ' AND t.date >= ?'
            params.append(start_date)

        if end_date:
            query += ' AND t.date <= ?'
            params.append(end_date)

        query += ' ORDER BY t.date DESC'

        with get_db_connection() as conn:
            transactions = conn.execute(query, params).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        headers = ['Дата', 'Тип', 'Сумма', 'Категория', 'Описание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for row, trans in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=trans['date'])
            ws.cell(row=row, column=2, value='Доход' if trans['type'] == 'income' else 'Расход')
            ws.cell(row=row, column=3, value=trans['amount'])
            ws.cell(row=row, column=4, value=trans['category'] or '')
            ws.cell(row=row, column=5, value=trans['description'] or '')

        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name='transactions.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/quick_add', methods=['POST'])
    @login_required
    def quick_add():
        """Быстрое добавление транзакции"""
        data = request.json
        type_ = data.get('type')
        amount = data.get('amount')

        if type_ and amount:
            with get_db_connection() as conn:
                if type_ == 'income':
                    category = conn.execute(
                        'SELECT id FROM categories WHERE type = "income" LIMIT 1'
                    ).fetchone()
                else:
                    category = conn.execute(
                        'SELECT id FROM categories WHERE type = "expense" LIMIT 1'
                    ).fetchone()

                category_id = category['id'] if category else None

                conn.execute('''
                    INSERT INTO transactions (type, amount, category_id, date, user_id)
                    VALUES (?, ?, ?, ?, ?)
                ''', (type_, amount, category_id, datetime.now().strftime('%Y-%m-%d'), session['user_id']))
                conn.commit()

            return jsonify({'success': True})

        return jsonify({'success': False, 'error': 'Неверные данные'})